"""
Tabilo — timetable grid data assembly + PDF/Excel rendering (Phase 10)

build_grid_data(institution, term, class_division) is the exact data
TimetableGridView has returned as JSON since Phase 5b, pulled out here so
the new PDF/Excel export endpoints render from the identical query this
view uses -- never a second, driftable copy of "resolve the grid." Keep
this function's return shape in lockstep with what the frontend's
GET /api/timetable-grid/ consumers expect; core/api/views.py's
TimetableGridView just wraps this dict in a Response() directly.

render_timetable_pdf()/render_timetable_xlsx() both pivot that same dict
into a day x period table (mirroring frontend/src/utils/timetableGrid.js's
pivotSlotsToGrid()) and format it for a specific file type. Both are pure
functions: dict in, bytes out -- no request/response concerns, so
core/api/views.py's export views stay thin wrappers.
"""

import io
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from core.models import TimeSlot, TimetableCell

SCHOOL_DAY_LABELS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

HEADER_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
BREAK_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")


def build_grid_data(institution, term, class_division):
    slots = TimeSlot.objects.filter(institution=institution).order_by("day_identifier", "period_number")

    # Avoid N+1: one query for every relevant TimetableCell, keyed by
    # time_slot_id, then attach in-memory while iterating slots below --
    # not a per-slot query.
    cells_by_slot_id = {
        cell.time_slot_id: cell
        for cell in TimetableCell.objects.filter(
            term=term, class_division=class_division
        ).select_related("subject", "teacher")
    }

    slots_payload = []
    for slot in slots:
        cell = cells_by_slot_id.get(slot.id)
        slots_payload.append({
            "id": slot.id,
            "day_identifier": slot.day_identifier,
            "period_number": slot.period_number,
            "start_time": slot.start_time,
            "end_time": slot.end_time,
            "is_break": slot.is_break,
            "cell": {
                "subject": cell.subject.name,
                "teacher": cell.teacher.name,
                "is_locked": cell.is_locked,
                "elective_group_id": cell.elective_group_id,
            } if cell is not None else None,
        })

    return {
        "institution": {
            "id": institution.id,
            "name": institution.name,
            "institution_type": institution.institution_type,
            "cycle_length": institution.cycle_length,
        },
        "term": {"id": term.id, "name": term.name, "is_active": term.is_active},
        "class_division": {
            "id": class_division.id, "name": class_division.name, "section": class_division.section,
        },
        "slots": slots_payload,
    }


def export_filename(grid_data, extension):
    class_division_part = _slugify(
        f"{grid_data['class_division']['name']}-{grid_data['class_division']['section']}"
    )
    term_part = _slugify(grid_data["term"]["name"])
    return f"{class_division_part}-{term_part}-timetable.{extension}"


def render_timetable_pdf(grid_data):
    day_identifiers, period_numbers, slot_by_day_and_period = _pivot_grid_data(grid_data)
    institution_type = grid_data["institution"]["institution_type"]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    cell_style = styles["Normal"].clone("TimetableCell")
    cell_style.fontSize = 8
    cell_style.leading = 10
    cell_style.alignment = 1  # TA_CENTER
    header_style = cell_style.clone("TimetableHeader")
    header_style.fontName = "Helvetica-Bold"
    break_style = cell_style.clone("TimetableBreakCell")
    break_style.fontName = "Helvetica-Oblique"

    # Paragraph parses its text as mini-XML -- every piece of user-editable
    # text (institution/term/class-division names, subject/teacher names)
    # goes through xml_escape() before being handed to a Paragraph, so a
    # stray "&" or "<" in any of them can't break the parser. Plain Table
    # cells (not wrapped in Paragraph) don't reliably wrap or line-break
    # embedded text, so every cell here is a Paragraph, not a bare string.
    title = (
        f"{xml_escape(grid_data['institution']['name'])} — "
        f"{xml_escape(grid_data['class_division']['name'])} - "
        f"{xml_escape(grid_data['class_division']['section'])} — "
        f"{xml_escape(grid_data['term']['name'])}"
    )

    header_row = [Paragraph("Period", header_style)] + [
        Paragraph(xml_escape(_day_label(institution_type, day)), header_style) for day in day_identifiers
    ]
    table_rows = [header_row]
    break_row_indices = set()

    for period_number in period_numbers:
        representative_slot = _representative_slot(slot_by_day_and_period, day_identifiers, period_number)
        is_break = representative_slot["is_break"] if representative_slot else False

        row_style = break_style if is_break else cell_style
        period_label = f"P{period_number}"
        if representative_slot:
            period_label += (
                f"<br/>{_format_time(representative_slot['start_time'])}"
                f"-{_format_time(representative_slot['end_time'])}"
            )

        row = [Paragraph(period_label, row_style)]
        for day in day_identifiers:
            text = _cell_text(slot_by_day_and_period.get((day, period_number)), separator="<br/>", escape=True)
            row.append(Paragraph(text, row_style))
        if is_break:
            break_row_indices.add(len(table_rows))
        table_rows.append(row)

    table = Table(table_rows, repeatRows=1)
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for row_index in break_row_indices:
        style_commands.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#e5e7eb")))
    table.setStyle(TableStyle(style_commands))

    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12), table]
    doc.build(elements)
    return buffer.getvalue()


def render_timetable_xlsx(grid_data):
    day_identifiers, period_numbers, slot_by_day_and_period = _pivot_grid_data(grid_data)
    institution_type = grid_data["institution"]["institution_type"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Timetable"

    title = (
        f"{grid_data['institution']['name']} - "
        f"{grid_data['class_division']['name']} - {grid_data['class_division']['section']} - "
        f"{grid_data['term']['name']}"
    )
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(day_identifiers) + 1)
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)

    header_row_num = 2
    header_period_cell = sheet.cell(row=header_row_num, column=1, value="Period")
    header_period_cell.font = Font(bold=True)
    header_period_cell.fill = HEADER_FILL
    for col, day in enumerate(day_identifiers, start=2):
        cell = sheet.cell(row=header_row_num, column=col, value=_day_label(institution_type, day))
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_offset, period_number in enumerate(period_numbers, start=1):
        row_num = header_row_num + row_offset
        representative_slot = _representative_slot(slot_by_day_and_period, day_identifiers, period_number)
        is_break = representative_slot["is_break"] if representative_slot else False

        period_label = f"P{period_number}"
        if representative_slot:
            period_label += (
                f" {_format_time(representative_slot['start_time'])}"
                f"-{_format_time(representative_slot['end_time'])}"
            )
        period_cell = sheet.cell(row=row_num, column=1, value=period_label)
        period_cell.font = Font(bold=True)
        if is_break:
            period_cell.fill = BREAK_FILL

        for col, day in enumerate(day_identifiers, start=2):
            value = _cell_text(slot_by_day_and_period.get((day, period_number)), separator=" - ")
            cell = sheet.cell(row=row_num, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if is_break:
                cell.fill = BREAK_FILL

    sheet.column_dimensions[get_column_letter(1)].width = 16
    for col_idx in range(2, len(day_identifiers) + 2):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 22

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _day_label(institution_type, day_identifier):
    if institution_type == "SCHOOL":
        index = day_identifier - 1
        if 0 <= index < len(SCHOOL_DAY_LABELS):
            return SCHOOL_DAY_LABELS[index]
    return f"Day {day_identifier}"


def _pivot_grid_data(grid_data):
    """Same pivot frontend/src/utils/timetableGrid.js's pivotSlotsToGrid()
    does, so the exported tables match the on-screen grid's shape."""
    slots = grid_data["slots"]
    day_identifiers = sorted({slot["day_identifier"] for slot in slots})
    period_numbers = sorted({slot["period_number"] for slot in slots})
    slot_by_day_and_period = {(slot["day_identifier"], slot["period_number"]): slot for slot in slots}
    return day_identifiers, period_numbers, slot_by_day_and_period


def _representative_slot(slot_by_day_and_period, day_identifiers, period_number):
    """A period's start/end time and is_break flag are consistent across
    every day (generate_time_slots() repeats the same daily template) --
    the first day is a safe representative for the row label."""
    if not day_identifiers:
        return None
    return slot_by_day_and_period.get((day_identifiers[0], period_number))


def _cell_text(slot, separator, escape=False):
    if slot is None:
        return ""
    if slot["is_break"]:
        return "Break"
    if slot["cell"] is None:
        return ""
    subject, teacher = slot["cell"]["subject"], slot["cell"]["teacher"]
    if escape:
        subject, teacher = xml_escape(subject), xml_escape(teacher)
    return f"{subject}{separator}{teacher}"


def _format_time(value):
    return value.strftime("%H:%M")


def _slugify(text):
    import re
    return re.sub(r"[^A-Za-z0-9]+", "-", text).strip("-") or "export"
