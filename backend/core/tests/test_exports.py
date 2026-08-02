import datetime
import io

from openpyxl import load_workbook
from pypdf import PdfReader
from rest_framework import status

from core.models import ClassDivision, CourseRequirement, Department, Subject, Teacher, UserProfile
from core.solver.apply import apply_solution
from core.solver.build import solve_feasibility
from core.tests.test_api import ApiTestCase

PDF_CONTENT_TYPE = "application/pdf"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ExportTestCase(ApiTestCase):
    """A solved grid to export -- institution/term/class-division/subject/
    teacher plus a real CP-SAT solve, same fixture-building style as
    core/tests/test_api.py's TimetableGridContentTests."""

    def setUp(self):
        self.institution_a = self.make_institution("Institution A")
        self.slots = self.make_grid(self.institution_a, periods_per_day=4)
        self.term_a = self.make_term(self.institution_a, "Term A")
        self.department_a = Department.objects.create(institution=self.institution_a, name="Dept A")
        self.division_a = ClassDivision.objects.create(
            institution=self.institution_a, department=self.department_a, name="Div A", section="A"
        )
        self.subject_a = Subject.objects.create(
            institution=self.institution_a, department=self.department_a, name="Math", code="MATH1"
        )
        self.teacher_a = Teacher.objects.create(
            institution=self.institution_a, name="Teacher One", email="teacher.one@test.edu",
            max_periods_per_day=10, max_periods_per_week=40,
        )
        CourseRequirement.objects.create(
            institution=self.institution_a, term=self.term_a, class_division=self.division_a,
            subject=self.subject_a, teacher=self.teacher_a, periods_per_week=2,
        )
        result = solve_feasibility(self.term_a)
        self.assertEqual(result.status, "FEASIBLE")
        apply_solution(self.term_a, result.assignments)

        self.institution_b = self.make_institution("Institution B")
        self.make_grid(self.institution_b)
        self.term_b = self.make_term(self.institution_b, "Term B")
        department_b = Department.objects.create(institution=self.institution_b, name="Dept B")
        self.division_b = ClassDivision.objects.create(
            institution=self.institution_b, department=department_b, name="Div B", section="A"
        )

        self.user_a = self.make_user(self.institution_a, "user_a", role=UserProfile.TEACHER)


class PdfExportTests(ExportTestCase):
    url = "/api/exports/timetable.pdf"

    def test_missing_params_returns_400(self):
        self.authenticate(self.user_a)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_foreign_term_id_returns_404(self):
        self.authenticate(self.user_a)
        response = self.client.get(f"{self.url}?term_id={self.term_b.id}&class_division_id={self.division_a.id}")
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_foreign_class_division_id_returns_404(self):
        self.authenticate(self.user_a)
        response = self.client.get(f"{self.url}?term_id={self.term_a.id}&class_division_id={self.division_b.id}")
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_teacher_role_can_export_own_class(self):
        """Deliberately NOT gated to ADMIN/COORDINATOR -- exporting a
        printed schedule is reading data the role can already see on
        screen via GET /api/timetable-grid/, not a management action."""
        self.authenticate(self.user_a)
        response = self.client.get(f"{self.url}?term_id={self.term_a.id}&class_division_id={self.division_a.id}")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_requires_authentication(self):
        response = self.client.get(f"{self.url}?term_id={self.term_a.id}&class_division_id={self.division_a.id}")
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_returns_a_real_readable_pdf(self):
        self.authenticate(self.user_a)
        response = self.client.get(f"{self.url}?term_id={self.term_a.id}&class_division_id={self.division_a.id}")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], PDF_CONTENT_TYPE)
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".pdf", response["Content-Disposition"])

        body = response.content
        self.assertTrue(len(body) > 0)
        self.assertTrue(body.startswith(b"%PDF"))

        reader = PdfReader(io.BytesIO(body))
        self.assertGreaterEqual(len(reader.pages), 1)
        text = reader.pages[0].extract_text()
        self.assertIn("Institution A", text)
        self.assertIn("Term A", text)
        self.assertIn("Div A", text)
        self.assertIn("Math", text)
        self.assertIn("Teacher One", text)
        self.assertIn("Break", text)


class ExcelExportTests(ExportTestCase):
    url = "/api/exports/timetable.xlsx"

    def test_missing_params_returns_400(self):
        self.authenticate(self.user_a)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_foreign_term_id_returns_404(self):
        self.authenticate(self.user_a)
        response = self.client.get(f"{self.url}?term_id={self.term_b.id}&class_division_id={self.division_a.id}")
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_foreign_class_division_id_returns_404(self):
        self.authenticate(self.user_a)
        response = self.client.get(f"{self.url}?term_id={self.term_a.id}&class_division_id={self.division_b.id}")
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_teacher_role_can_export_own_class(self):
        self.authenticate(self.user_a)
        response = self.client.get(f"{self.url}?term_id={self.term_a.id}&class_division_id={self.division_a.id}")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_requires_authentication(self):
        response = self.client.get(f"{self.url}?term_id={self.term_a.id}&class_division_id={self.division_a.id}")
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_returns_a_real_readable_xlsx(self):
        self.authenticate(self.user_a)
        response = self.client.get(f"{self.url}?term_id={self.term_a.id}&class_division_id={self.division_a.id}")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], XLSX_CONTENT_TYPE)
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

        body = response.content
        self.assertTrue(len(body) > 0)
        self.assertTrue(body.startswith(b"PK"))  # xlsx is a zip container

        workbook = load_workbook(io.BytesIO(body))
        sheet = workbook.active
        self.assertEqual(sheet.title, "Timetable")

        all_values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value]
        self.assertTrue(any("Institution A" in str(v) for v in all_values))
        self.assertTrue(any("Term A" in str(v) for v in all_values))
        self.assertTrue(any(v == "Math - Teacher One" for v in all_values))
        self.assertTrue(any(v == "Break" for v in all_values))
